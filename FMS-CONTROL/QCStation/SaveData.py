# -*- coding: utf-8 -*-
"""
Created on Fri Mar 23 12:11:12 2018

@author: cap
"""

from openpyxl import Workbook as WB
from openpyxl import load_workbook
import datetime
from random import randint
now = datetime.datetime.now()

class SaveData():

    def __init__(self):
         self.wb = WB()
         self.ws = self.wb.active
         self.a = 1
         self.name = str(now.year)+"-"+str(now.month)+"-"+str(now.day)+"-"+str(now.hour)+"-"+str(now.minute)+"-"+str(now.second)+"--"+str(randint(111, 1000))

    def RecordData(self,D1,D2):
        ws = self.ws 
        b = self.a   
        ws.cell(row=b, column=1,value = b)
        ws.cell(row=b, column=2,value = D1)
        ws.cell(row=b, column=3,value = D2)
        self.a = b+1
        
    def SaveFile(self):
        self.wb.save(self.name + '.xlsx')
        print("data save in file")
        
        